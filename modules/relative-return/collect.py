# -*- coding: utf-8 -*-
r"""
상대수익률 추이 수집 모듈 (lighthouse-amc-dashboard)

WISE 업종지수 시계열 엑셀(상대수익률차트.xlsx)을 읽어, 각 KSE 업종의
'KOSPI 대비' 1개월/3개월 수익률을 최근일 기준 5개 시점(6개월전·3개월전·
1개월전·1주전·최근)으로 뽑아 data/relative-return.json 스냅샷을 만든다.

- 상대수익률 = 업종 수익률 − 코스피 수익률 (같은 기간)
- X축: KOSPI 대비 1개월 수익률, Y축: KOSPI 대비 3개월 수익률
- 최근일 방향으로 화살표를 그리기 위해 시점 순서는 과거→최근으로 정렬한다.

엑셀 경로는 REL_RETURN_XLSX 환경변수로 바꿀 수 있다.
사용: python modules/relative-return/collect.py
"""
import json
import os
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
OUTPUT_PATH = REPO_ROOT / "data" / "relative-return.json"
XLSX_PATH = Path(os.environ.get(
    "REL_RETURN_XLSX",
    r"C:\Users\CHECK\Documents\Data\상대수익률차트.xlsx",
))

CODE_ROW, NAME_ROW, DATA_START = 8, 9, 15
KOSPI_CODE = "IKS900"
KOSDAQ_CODE = "IKQ900"

# 최근일 기준 되돌아볼 시점 (라벨, 개월수 or 일수)
# 차트는 1주 전 → 최근 두 시점만 사용한다(6/3/1개월 제외).
POINTS = [
    ("1w", "1주 전", {"days": 7}),
    ("now", "최근", {"days": 0}),
]


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _add_offset(d: date, off: dict) -> date:
    if "days" in off:
        return d - timedelta(days=off["days"])
    # months
    m = off["months"]
    y, mo = d.year, d.month - m
    while mo <= 0:
        mo += 12
        y -= 1
    day = min(d.day, 28)
    return date(y, mo, day)


def _nearest_row(dates: list, target: date) -> int:
    """target 이하에서 가장 가까운 인덱스. 없으면 전체 중 최근접."""
    best_le, best_le_i = None, None
    best_any, best_any_i = None, None
    for i, d in enumerate(dates):
        if d is None:
            continue
        diff = abs((d - target).days)
        if best_any is None or diff < best_any:
            best_any, best_any_i = diff, i
        if d <= target and (best_le is None or (target - d).days < best_le):
            best_le, best_le_i = (target - d).days, i
    return best_le_i if best_le_i is not None else best_any_i


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"[!] 엑셀을 찾을 수 없습니다: {XLSX_PATH}")
        return 1

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb.active
    ncol = ws.max_column

    codes = [ws.cell(row=CODE_ROW, column=c).value for c in range(1, ncol + 1)]
    names = [ws.cell(row=NAME_ROW, column=c).value for c in range(1, ncol + 1)]

    # 지수별 (1개월 컬럼, 3개월 컬럼) 매핑 — 각 지수는 연속 2컬럼(1M, 3M)
    index_cols = {}   # code -> {"name", "c1m", "c3m"}
    c = 2
    while c <= ncol:
        code = codes[c - 1]
        if code:
            index_cols[code] = {"name": names[c - 1], "c1m": c, "c3m": c + 1}
        c += 2

    if KOSPI_CODE not in index_cols:
        print("[!] 코스피(IKS900) 컬럼을 찾지 못했습니다.")
        return 1
    k1m = index_cols[KOSPI_CODE]["c1m"]
    k3m = index_cols[KOSPI_CODE]["c3m"]

    # 날짜 목록
    dates = []
    for r in range(DATA_START, ws.max_row + 1):
        d = _to_date(ws.cell(row=r, column=1).value)
        dates.append(d)
    valid_idx = [i for i, d in enumerate(dates) if d is not None]
    if not valid_idx:
        print("[!] 날짜 데이터를 찾지 못했습니다.")
        return 1
    base_date = dates[valid_idx[-1]]

    # 5개 시점의 행 인덱스(=DATA_START 기준 오프셋)
    picked = []
    for key, label, off in POINTS:
        target = _add_offset(base_date, off)
        i = _nearest_row(dates, target)
        picked.append({"key": key, "label": label,
                       "date": dates[i].isoformat(), "row": DATA_START + i})

    def val(row, col):
        v = ws.cell(row=row, column=col).value
        return float(v) if isinstance(v, (int, float)) else None

    sectors = []
    for code, info in index_cols.items():
        if code in (KOSPI_CODE, KOSDAQ_CODE):
            continue
        traj = []
        for p in picked:
            r = p["row"]
            s1, s3 = val(r, info["c1m"]), val(r, info["c3m"])
            b1, b3 = val(r, k1m), val(r, k3m)
            if None in (s1, s3, b1, b3):
                traj.append(None)
            else:
                traj.append([round(s1 - b1, 2), round(s3 - b3, 2)])
        name = (info["name"] or code).replace("KSE ", "").strip()
        sectors.append({"code": code, "name": name, "traj": traj})

    payload = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S%z"),
        "base_date": base_date.isoformat(),
        "x_label": "(KOSPI 대비 1개월 수익률)",
        "y_label": "(KOSPI 대비 3개월 수익률)",
        "points": [{"key": p["key"], "label": p["label"], "date": p["date"]} for p in picked],
        "sectors": sectors,
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"[OK] 업종 {len(sectors)}개 → {OUTPUT_PATH}")
    print(f"     기준일: {base_date}  시점: " + ", ".join(f"{p['label']}={p['date']}" for p in picked))
    for s in sectors:
        if s["name"] in ("반도체", "IT하드웨어", "소매(유통)", "에너지"):
            print(f"     {s['name']}: 최근={s['traj'][-1]}  1주전={s['traj'][0]}")
    return 0


if __name__ == "__main__":
    import sys
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    raise SystemExit(main())
